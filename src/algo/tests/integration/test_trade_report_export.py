"""End-to-end proof that closing a position through the real wired stack
(DependencyContainer -> Strategy1 -> ExitLogic) triggers the trade-report
export, and that the produced .xlsx matches the trades the strategy actually
persisted.

The export is opt-in via app.yaml's ``report_output_dir``; the rest of the
integration suite leaves it unset, so only this file exercises the exporter
through the container. See test_trade_report.py for the exporter's own
column/format unit coverage.
"""

from __future__ import annotations

from decimal import Decimal

from openpyxl import load_workbook

from algo.brokers.simulation import StaticPriceSource
from algo.common.enums import ExitReason, OptionType, PositionState
from algo.database.repositories.position_repository import PositionRepository
from algo.reporting.trade_report import format_instrument
from algo.tests.integration.conftest import (
    atm_legs,
    build_container,
    build_nifty_option_chain,
    make_clock,
)


def _entry(container, runner, clock):
    clock.set_time(hour=9, minute=20)
    runner.dispatch_time_trigger("entry")
    instance_id = runner._strategy.context.identity.instance_id  # noqa: SLF001
    with container.session_factory() as session:
        return PositionRepository(session).get_by_instance_and_date(instance_id, clock.today())


class TestExportFiresOnClose:
    def test_closed_position_writes_the_daily_report(self, tmp_path):
        clock = make_clock(hour=9, minute=0)
        catalog = build_nifty_option_chain()
        call, put = atm_legs(catalog)
        prices = StaticPriceSource({call: Decimal("100"), put: Decimal("100")})
        reports_dir = tmp_path / "out"
        container = build_container(
            tmp_path, clock=clock, db_path=tmp_path / "db.sqlite",
            instrument_catalog=catalog, price_source=prices,
            # POSIX-style so the path sits cleanly inside a double-quoted YAML
            # scalar on Windows (backslashes would read as escape sequences).
            report_output_dir=reports_dir.as_posix(),
        )
        container.start()
        try:
            runner = container.runners[0]
            _entry(container, runner, clock)

            # Drop premium to target and force the exit through the real path.
            prices.set_price(call, Decimal("80"))
            prices.set_price(put, Decimal("77"))
            clock.set_time(hour=12, minute=30)
            runner.dispatch_time_trigger("cutoff")

            instance_id = runner._strategy.context.identity.instance_id  # noqa: SLF001
            with container.session_factory() as session:
                repo = PositionRepository(session)
                position = repo.get_by_instance_and_date(instance_id, clock.today())
                assert position.state is PositionState.CLOSED
                assert position.exit_reason is ExitReason.TARGET
                legs = {t.option_type: t for t in repo.list_trades_for_position(position.id)}
                expected = {
                    opt: format_instrument(
                        exchange=leg.exchange.value,
                        underlying="NIFTY",
                        expiry=position.expiry_date,
                        strike=leg.strike,
                        option_type=opt,
                    )
                    for opt, leg in legs.items()
                }
                ce_leg, pe_leg = legs[OptionType.CE], legs[OptionType.PE]
        finally:
            container.stop()

        # The report exists for the trading day and reflects both legs.
        path = reports_dir / f"trades_{clock.today():%d-%m-%Y}.xlsx"
        assert path.exists()
        rows = [list(r) for r in load_workbook(path).active.iter_rows(values_only=True)]

        from algo.reporting.trade_report import _COLUMNS

        assert rows[0] == list(_COLUMNS)  # 12 columns incl. the trade-level P&L ones
        assert len(rows) == 3  # header + CE + PE
        # The trade-level columns are present and shared by both legs.
        assert rows[1][7] == rows[2][7] == "TARGET"  # Exit Reason (target hit)
        assert rows[1][11] == rows[2][11]  # Net P&L identical on both legs of a trade
        assert [r[0] for r in rows[1:]] == [1, 2]
        assert rows[1][2] == expected[OptionType.CE]
        assert rows[2][2] == expected[OptionType.PE]
        # Prices are the executed averages the strategy stored, to 2 decimals.
        assert rows[1][4] == float(ce_leg.entry_price)
        assert rows[1][6] == float(ce_leg.exit_price)
        assert rows[2][4] == float(pe_leg.entry_price)
        assert rows[2][6] == float(pe_leg.exit_price)
        # Times are HH:MM:SS strings.
        for r in rows[1:]:
            assert len(r[3].split(":")) == 3
            assert len(r[5].split(":")) == 3

    def test_export_disabled_by_default_writes_nothing(self, tmp_path):
        """With report_output_dir unset (the suite-wide default), a close must
        not write any report -- the feature is strictly opt-in."""
        clock = make_clock(hour=9, minute=0)
        catalog = build_nifty_option_chain()
        call, put = atm_legs(catalog)
        prices = StaticPriceSource({call: Decimal("100"), put: Decimal("100")})
        container = build_container(
            tmp_path, clock=clock, db_path=tmp_path / "db.sqlite",
            instrument_catalog=catalog, price_source=prices,
        )
        container.start()
        try:
            runner = container.runners[0]
            _entry(container, runner, clock)
            prices.set_price(call, Decimal("80"))
            prices.set_price(put, Decimal("77"))
            clock.set_time(hour=12, minute=30)
            runner.dispatch_time_trigger("cutoff")
            assert container.trade_exporter is None
        finally:
            container.stop()

        assert not list(tmp_path.glob("**/trades_*.xlsx"))
