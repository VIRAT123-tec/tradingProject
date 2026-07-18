"""Tests for the per-leg trade report exporter (reporting/trade_report.py).

Builds real CLOSED positions/legs in in-memory SQLite (with the JSONB/
BigInteger shims the rest of the suite uses), runs the exporter, and reads the
produced .xlsx back with openpyxl to assert the exact columns, formats, and
one-row-per-leg contract the report requires.
"""

from __future__ import annotations

from datetime import date, datetime, timezone
from decimal import Decimal

from openpyxl import load_workbook
from sqlalchemy import BigInteger, create_engine
from sqlalchemy.dialects.postgresql import JSONB
from sqlalchemy.ext.compiler import compiles
from sqlalchemy.orm import sessionmaker


@compiles(BigInteger, "sqlite")
def _bigint_sqlite(type_, compiler, **kw):  # noqa: ANN001, ANN202
    return "INTEGER"


@compiles(JSONB, "sqlite")
def _jsonb_sqlite(type_, compiler, **kw):  # noqa: ANN001, ANN202
    return "JSON"


from algo.common.enums import (
    Exchange,
    ExitReason,
    OptionType,
    PositionState,
    TradeLegStatus,
)
from algo.database.models import Account, Base, Position, StrategyInstance, Trade
from algo.reporting.trade_report import _COLUMNS, TradeReportExporter, format_instrument

TRADE_DATE = date(2026, 7, 13)
EXPIRY = date(2026, 7, 14)
# 09:20:00 and 12:30:00 IST expressed as the UTC that is actually stored
# (IST = UTC + 5:30): 03:50 UTC == 09:20 IST, 07:00 UTC == 12:30 IST.
ENTRY_UTC = datetime(2026, 7, 13, 3, 50, 0, tzinfo=timezone.utc)
EXIT_UTC = datetime(2026, 7, 13, 7, 0, 0, tzinfo=timezone.utc)


def _factory():
    engine = create_engine("sqlite://")
    Base.metadata.create_all(engine)
    return sessionmaker(bind=engine, expire_on_commit=False)


def _make_closed_position(
    session_factory,
    *,
    instrument: str,
    exchange: Exchange,
    strike: Decimal,
    ce_entry: Decimal,
    ce_exit: Decimal,
    pe_entry: Decimal,
    pe_exit: Decimal,
    expiry: date = EXPIRY,
    state: PositionState = PositionState.CLOSED,
    exit_completed_at: datetime | None = EXIT_UTC,
    realized_pnl: Decimal | None = None,
    exit_reason: ExitReason = ExitReason.TARGET,
    trade_date: date = TRADE_DATE,
) -> int:
    """Persist a (StrategyInstance, Position, CE leg, PE leg) graph and return
    the position id. Legs carry the SELL entry / BUY exit averages and the
    UTC fill timestamps exactly as the strategy would have written them."""
    with session_factory() as s:
        account = Account(broker="SIMULATION", display_name="test")
        s.add(account)
        s.flush()
        instance = StrategyInstance(
            strategy_id="strategy_1", instrument=instrument, account_id=account.id, exchange=exchange
        )
        s.add(instance)
        s.flush()
        position = Position(
            strategy_instance_id=instance.id,
            trade_date=trade_date,
            state=state,
            expiry_date=expiry,
            strike=strike,
            quantity=75,
            exit_reason=exit_reason,
            realized_pnl=realized_pnl,
            exit_completed_at=exit_completed_at,
        )
        s.add(position)
        s.flush()
        for opt, entry, exit_ in (
            (OptionType.CE, ce_entry, ce_exit),
            (OptionType.PE, pe_entry, pe_exit),
        ):
            s.add(
                Trade(
                    position_id=position.id,
                    option_type=opt,
                    trading_symbol=f"{instrument}SIM{opt.value}",  # broker-format, deliberately NOT what we export
                    exchange=exchange,
                    strike=strike,
                    quantity=75,
                    entry_price=entry,
                    exit_price=exit_,
                    entry_time=ENTRY_UTC,
                    exit_time=EXIT_UTC,
                    status=TradeLegStatus.CLOSED,
                )
            )
        s.commit()
        return position.id


def _read_rows(path):
    wb = load_workbook(path)
    ws = wb.active
    return [list(row) for row in ws.iter_rows(values_only=True)]


class TestInstrumentFormat:
    def test_nifty_ce_matches_required_pattern(self):
        assert (
            format_instrument(
                exchange="NFO", underlying="NIFTY", expiry=date(2026, 7, 14),
                strike=Decimal("24200.00"), option_type=OptionType.CE,
            )
            == "NFO:NIFTY26071424200CE"
        )

    def test_sensex_pe_matches_required_pattern(self):
        assert (
            format_instrument(
                exchange="BFO", underlying="SENSEX", expiry=date(2026, 7, 16),
                strike=Decimal("77600.00"), option_type=OptionType.PE,
            )
            == "BFO:SENSEX26071677600PE"
        )


class TestExportDay:
    def test_header_and_one_row_per_leg(self, tmp_path):
        sf = _factory()
        _make_closed_position(
            sf, instrument="NIFTY", exchange=Exchange.NFO, strike=Decimal("24200"),
            ce_entry=Decimal("80"), ce_exit=Decimal("77"),
            pe_entry=Decimal("100"), pe_exit=Decimal("85"),
            realized_pnl=Decimal("1350"),
        )
        exporter = TradeReportExporter(session_factory=sf, output_dir=tmp_path)

        path = exporter.export_day(TRADE_DATE)

        assert path == tmp_path / "trades_13-07-2026.xlsx"
        rows = _read_rows(path)
        assert rows[0] == list(_COLUMNS)  # 12 columns incl. the new trade-level ones
        # Two legs -> two data rows, CE before PE; the last five columns are
        # trade-level so both legs carry the same values.
        assert len(rows) == 3
        assert rows[1] == [
            1, "13-07-2026", "NFO:NIFTY26071424200CE", "09:20:00", 80.00, "12:30:00", 77.00,
            "TARGET", 1350.00, 1350.00, 0.00, 1350.00,
        ]
        assert rows[2] == [
            2, "13-07-2026", "NFO:NIFTY26071424200PE", "09:20:00", 100.00, "12:30:00", 85.00,
            "TARGET", 1350.00, 1350.00, 0.00, 1350.00,
        ]

    def test_prices_render_with_two_decimals(self, tmp_path):
        sf = _factory()
        _make_closed_position(
            sf, instrument="NIFTY", exchange=Exchange.NFO, strike=Decimal("24200"),
            ce_entry=Decimal("80.35"), ce_exit=Decimal("77.05"),
            pe_entry=Decimal("100.5"), pe_exit=Decimal("85.1"),
        )
        exporter = TradeReportExporter(session_factory=sf, output_dir=tmp_path)
        path = exporter.export_day(TRADE_DATE)

        wb = load_workbook(path)
        ws = wb.active
        # Values preserved to 2dp and the cell is formatted to always show 2dp.
        assert ws.cell(row=2, column=5).value == 80.35
        assert ws.cell(row=2, column=5).number_format == "0.00"
        assert ws.cell(row=3, column=5).value == 100.50
        assert ws.cell(row=3, column=7).value == 85.10

    def test_multiple_instruments_accumulate_into_one_daily_file(self, tmp_path):
        sf = _factory()
        _make_closed_position(
            sf, instrument="NIFTY", exchange=Exchange.NFO, strike=Decimal("24200"),
            ce_entry=Decimal("80"), ce_exit=Decimal("77"),
            pe_entry=Decimal("100"), pe_exit=Decimal("85"),
            exit_completed_at=datetime(2026, 7, 13, 6, 0, tzinfo=timezone.utc),
        )
        _make_closed_position(
            sf, instrument="SENSEX", exchange=Exchange.BFO, strike=Decimal("77600"),
            ce_entry=Decimal("450"), ce_exit=Decimal("400"),
            pe_entry=Decimal("470"), pe_exit=Decimal("420"),
            expiry=date(2026, 7, 16),  # Sensex weekly expiry differs from Nifty's
            exit_completed_at=datetime(2026, 7, 13, 7, 0, tzinfo=timezone.utc),
        )
        exporter = TradeReportExporter(session_factory=sf, output_dir=tmp_path)
        rows = _read_rows(exporter.export_day(TRADE_DATE))

        # 4 legs, S.no 1..4, NIFTY (earlier exit) before SENSEX, CE before PE.
        assert [r[0] for r in rows[1:]] == [1, 2, 3, 4]
        assert [r[2] for r in rows[1:]] == [
            "NFO:NIFTY26071424200CE",
            "NFO:NIFTY26071424200PE",
            "BFO:SENSEX26071677600CE",
            "BFO:SENSEX26071677600PE",
        ]

    def test_rebuild_is_idempotent(self, tmp_path):
        sf = _factory()
        _make_closed_position(
            sf, instrument="NIFTY", exchange=Exchange.NFO, strike=Decimal("24200"),
            ce_entry=Decimal("80"), ce_exit=Decimal("77"),
            pe_entry=Decimal("100"), pe_exit=Decimal("85"),
        )
        exporter = TradeReportExporter(session_factory=sf, output_dir=tmp_path)
        first = _read_rows(exporter.export_day(TRADE_DATE))
        second = _read_rows(exporter.export_day(TRADE_DATE))  # re-export same day
        assert first == second

    def test_only_closed_positions_are_exported(self, tmp_path):
        sf = _factory()
        # An OPEN position for the same day must not appear in the report.
        _make_closed_position(
            sf, instrument="NIFTY", exchange=Exchange.NFO, strike=Decimal("24200"),
            ce_entry=Decimal("80"), ce_exit=Decimal("77"),
            pe_entry=Decimal("100"), pe_exit=Decimal("85"),
            state=PositionState.OPEN, exit_completed_at=None,
        )
        exporter = TradeReportExporter(session_factory=sf, output_dir=tmp_path)
        rows = _read_rows(exporter.export_day(TRADE_DATE))
        assert len(rows) == 1  # header only, no data rows


class TestCumulativeTotals:
    """The running Exit Reason / P&L / Total Profit / Total Loss / Net columns."""

    def _three_trades(self, sf, pnls):
        # Three trades on the same day, exit-ordered by the given P&L sequence.
        for i, pnl in enumerate(pnls, start=1):
            _make_closed_position(
                sf, instrument="NIFTY", exchange=Exchange.NFO, strike=Decimal("24200"),
                ce_entry=Decimal("100"), ce_exit=Decimal("90"),
                pe_entry=Decimal("100"), pe_exit=Decimal("90"),
                realized_pnl=Decimal(pnl),
                exit_completed_at=datetime(2026, 7, 13, 6, i, tzinfo=timezone.utc),
            )

    def test_profit_only_totals_never_decrease(self, tmp_path):
        sf = _factory()
        self._three_trades(sf, [500, -200, 300])  # user's Total Profit example
        rows = _read_rows(TradeReportExporter(session_factory=sf, output_dir=tmp_path).export_day(TRADE_DATE))
        # 3 trades x 2 legs; trade-level values repeat per pair.
        assert [r[9] for r in rows[1:]] == [500, 500, 500, 500, 800, 800]   # Total Profit

    def test_total_loss_is_cumulative_absolute(self, tmp_path):
        sf = _factory()
        self._three_trades(sf, [500, -200, -150])  # user's Total Loss example
        rows = _read_rows(TradeReportExporter(session_factory=sf, output_dir=tmp_path).export_day(TRADE_DATE))
        assert [r[10] for r in rows[1:]] == [0, 0, 200, 200, 350, 350]      # Total Loss (positive)

    def test_net_is_profit_minus_loss(self, tmp_path):
        sf = _factory()
        self._three_trades(sf, [500, -200, 100])  # user's Net example
        rows = _read_rows(TradeReportExporter(session_factory=sf, output_dir=tmp_path).export_day(TRADE_DATE))
        assert [r[8] for r in rows[1:]] == [500, 500, -200, -200, 100, 100]  # P&L (per trade)
        assert [r[11] for r in rows[1:]] == [500, 500, 300, 300, 400, 400]   # Net P&L

    def test_totals_continue_across_days_and_do_not_restart(self, tmp_path):
        sf = _factory()
        # A prior-day winning trade (+1000) and a prior-day losing trade (-300).
        _make_closed_position(
            sf, instrument="NIFTY", exchange=Exchange.NFO, strike=Decimal("24200"),
            ce_entry=Decimal("100"), ce_exit=Decimal("90"), pe_entry=Decimal("100"), pe_exit=Decimal("90"),
            realized_pnl=Decimal("1000"), trade_date=date(2026, 7, 10),
        )
        _make_closed_position(
            sf, instrument="NIFTY", exchange=Exchange.NFO, strike=Decimal("24200"),
            ce_entry=Decimal("100"), ce_exit=Decimal("90"), pe_entry=Decimal("100"), pe_exit=Decimal("90"),
            realized_pnl=Decimal("-300"), trade_date=date(2026, 7, 11),
        )
        # Today's trade (+200): totals must continue from the prior days.
        self._three_trades(sf, [200])
        rows = _read_rows(TradeReportExporter(session_factory=sf, output_dir=tmp_path).export_day(TRADE_DATE))
        # opening: profit 1000, loss 300 -> today +200 -> profit 1200, loss 300, net 900.
        assert rows[1][9] == 1200 and rows[1][10] == 300 and rows[1][11] == 900

    def test_exit_reason_column(self, tmp_path):
        sf = _factory()
        _make_closed_position(
            sf, instrument="NIFTY", exchange=Exchange.NFO, strike=Decimal("24200"),
            ce_entry=Decimal("100"), ce_exit=Decimal("120"), pe_entry=Decimal("100"), pe_exit=Decimal("120"),
            realized_pnl=Decimal("-3000"), exit_reason=ExitReason.STOPLOSS,
        )
        rows = _read_rows(TradeReportExporter(session_factory=sf, output_dir=tmp_path).export_day(TRADE_DATE))
        assert rows[1][7] == "STOPLOSS" and rows[2][7] == "STOPLOSS"


class TestExportClosedPosition:
    def test_resolves_day_from_position_and_writes_file(self, tmp_path):
        sf = _factory()
        pid = _make_closed_position(
            sf, instrument="NIFTY", exchange=Exchange.NFO, strike=Decimal("24200"),
            ce_entry=Decimal("80"), ce_exit=Decimal("77"),
            pe_entry=Decimal("100"), pe_exit=Decimal("85"),
        )
        exporter = TradeReportExporter(session_factory=sf, output_dir=tmp_path)

        exporter.export_closed_position(pid)

        assert (tmp_path / "trades_13-07-2026.xlsx").exists()

    def test_never_raises_on_unknown_position(self, tmp_path):
        sf = _factory()
        exporter = TradeReportExporter(session_factory=sf, output_dir=tmp_path)
        # Must be exception-safe: a missing position is logged and swallowed.
        exporter.export_closed_position(9999)
        assert not (tmp_path / "trades_13-07-2026.xlsx").exists()
